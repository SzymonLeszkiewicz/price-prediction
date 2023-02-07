import openpyxl
import csv

'''
getting data from excel file
'''

workbook = openpyxl.load_workbook('data/mieszkania.xlsx')
worksheet = workbook.active
# return values from a2 and a3
means = []

mean_list = [worksheet.cell(row=2, column=2).value, worksheet.cell(row=3, column=2).value]
mean = round(sum(mean_list) / len(mean_list), 2)
means.append(mean)

miasta = 'rok   Białystok	Bydgoszcz	Gdańsk	Gdynia	Katowice	Kielce	Kraków	Lublin	Łódź	Olsztyn	Opole	Poznań	Rzeszów	Szczecin	Warszawa	Wrocław	Zielona_Góra'.split()
# write to csv as header row
with open('data/srednie2.csv', 'w', encoding='utf-8', newline='') as csvplik:
    csvwriter = csv.writer(csvplik)
    csvwriter.writerow(miasta)

    years = [x for x in range(2006, 2022)]  # 2006-2021
    years = years[::-1]

    mean_row = [years.pop(-1)]

    for kol in range(2, worksheet.max_column + 1):

        srednia = (worksheet.cell(row=2, column=kol).value + worksheet.cell(row=3, column=kol).value)/2
        mean_row.append(round(srednia, 2))
    #save to csv
    csvwriter.writerow(mean_row)

    for i in range(4, worksheet.max_row, 4):
        mean_row = []
        mean_row.append(years.pop(-1))
        for kol in range(2, worksheet.max_column + 1):
            mean_list = []
            for j in range(4):
                mean_list.append(worksheet.cell(row=i + j, column=kol).value)
            mean = round(sum(mean_list) / len(mean_list), 2)
            means.append(mean)
            mean_row.append(mean)
        #sace mean_row to csv
        csvwriter.writerow(mean_row)

    mean_row = [2021]
    for kol in range(2, worksheet.max_column + 1):
        mean_row.append(round(worksheet.cell(row=60, column=kol).value,2))
    csvwriter.writerow(mean_row)




